"""
CMF Tarjetas — Scraper v2
Descarga datos de tarjetas emitidas directamente desde la API pública
de BEST CMF (best-cmf.cl) en formato Excel, sin necesidad de Playwright
ni Anthropic.

Series descargadas:
  - SBIF_TDEB_VIG_AGIFI_NUM          : Débito vigentes por emisor
  - SBIF_TCRED_BANC_VIGTIT_AGIFI_NUM : Crédito titulares por emisor
  - SBIF_TCRED_BANC_VIGADIC_AGIFI_NUM: Crédito adicionales por emisor
  - CMF_TPREP_NBANC_VIG_NAT_AGIFI_NUM_MONT: Prepago no bancario por institución
"""

import io
import json
import pathlib
import sys
import urllib.error
import urllib.request
from datetime import datetime

import openpyxl

# ─── Configuración ──────────────────────────────────────────────────────────

BEST_API = "https://best-sbif-api.azurewebsites.net"
BEST_EXCEL_URL = (
    BEST_API
    + "/CuadroExcel/?Tag={tag}&Orientacion=H&TodosLosElementos=true&nocache={ts}"
)

SERIES = {
    "debito":        "SBIF_TDEB_VIG_AGIFI_NUM",
    "cred_tit":      "SBIF_TCRED_BANC_VIGTIT_AGIFI_NUM",
    "cred_adic":     "SBIF_TCRED_BANC_VIGADIC_AGIFI_NUM",
    "prepago_nbanc": "CMF_TPREP_NBANC_VIG_NAT_AGIFI_NUM_MONT",
}

SCRIPT_DIR   = pathlib.Path(__file__).parent
REPO_ROOT    = SCRIPT_DIR.parent
OUTPUT_JSON  = REPO_ROOT / "frontend" / "src" / "data" / "latest.json"
LAST_PERIOD  = SCRIPT_DIR / "last_period.txt"

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 CMF-Tarjetas-Scraper/2.0",
    "Referer":    "https://www.best-cmf.cl/",
}

# Instituciones conocidas como cooperativas
COOPERATIVAS = {"coopeuch", "capual", "detacoop", "coocretal"}

# Instituciones conocidas como SAEs (emisores no bancarios de tarjetas de crédito)
# Reguladas bajo Ley 20.950 o como sociedades de apoyo al giro
SAES = {
    "cmr falabella",
    "cat administradora",
    "car s.a.",
    "servicios financieros y administración de créditos comerciales",
    "smu corp",
    "consorcio tarjetas de crédito",
}


# ─── Funciones de descarga ───────────────────────────────────────────────────

def download_excel(tag: str) -> openpyxl.Workbook:
    ts = int(datetime.utcnow().timestamp())
    url = BEST_EXCEL_URL.format(tag=tag, ts=ts)
    req = urllib.request.Request(url, headers=HTTP_HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return openpyxl.load_workbook(io.BytesIO(resp.read()))
    except urllib.error.HTTPError as e:
        print(f"ERROR HTTP {e.code} descargando {tag}", file=sys.stderr)
        raise


def parse_latest_row(wb: openpyxl.Workbook) -> tuple[datetime | None, dict[str, int]]:
    """
    Extrae la fila más reciente del Excel de BEST.
    Retorna (fecha, {institución: valor}).
    Los '-' se convierten a 0.
    """
    ws = wb.active

    # Fila 4 = cabeceras, fila 5+ = datos
    headers = [cell.value for cell in ws[4]]

    last_date: datetime | None = None
    last_values: list = []

    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        last_date = row[0]
        last_values = list(row)

    result: dict[str, int] = {}
    for col, val in zip(headers[1:], last_values[1:]):
        if col is None:
            continue
        name = col.strip()
        if val is None or val == "-":
            result[name] = 0
        else:
            try:
                result[name] = int(str(val).replace(",", "").replace(".", ""))
            except ValueError:
                result[name] = 0

    return last_date, result


def clean_name(name: str) -> str:
    """Normaliza el nombre de la institución para comparación."""
    return (
        name.lower()
        .replace("banco del estado de chile", "bancoestado")
        .replace("banco itaú chile (*)", "banco itaú chile")
        .strip()
    )


def tipo_for(name: str, is_prepago_nbanc: bool) -> str:
    normalized = clean_name(name)
    if any(coop in normalized for coop in COOPERATIVAS):
        return "cooperativa"
    if is_prepago_nbanc or any(sae in normalized for sae in SAES):
        return "no_bancario"
    return "banco"


def format_periodo(dt: datetime) -> str:
    meses = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    return f"{meses[dt.month - 1]} {dt.year}"


# ─── Lógica principal ────────────────────────────────────────────────────────

def main() -> None:
    print("Descargando series de BEST CMF...")

    # 1. Descargar todos los Excel
    wb_deb  = download_excel(SERIES["debito"])
    wb_ctit = download_excel(SERIES["cred_tit"])
    wb_cad  = download_excel(SERIES["cred_adic"])
    wb_prep = download_excel(SERIES["prepago_nbanc"])

    # 2. Parsear última fila de cada serie
    fecha_deb,  deb   = parse_latest_row(wb_deb)
    fecha_ctit, ctit  = parse_latest_row(wb_ctit)
    fecha_cad,  cadic = parse_latest_row(wb_cad)
    fecha_prep, prep  = parse_latest_row(wb_prep)

    print(f"  Débito:         {fecha_deb}  ({len(deb)} inst.)")
    print(f"  Crédito tit.:   {fecha_ctit}  ({len(ctit)} inst.)")
    print(f"  Crédito adic.:  {fecha_cad}  ({len(cadic)} inst.)")
    print(f"  Prepago nbanc:  {fecha_prep}  ({len(prep)} inst.)")

    # Usar la fecha más reciente disponible para el período
    fechas = [f for f in [fecha_deb, fecha_ctit, fecha_cad, fecha_prep] if f]
    periodo_dt = max(fechas) if fechas else datetime.utcnow()
    periodo_str = format_periodo(periodo_dt)

    # 3. Verificar si ya procesamos este período
    periodo_key = periodo_dt.strftime("%Y-%m")
    if LAST_PERIOD.exists():
        last = LAST_PERIOD.read_text().strip()
        if last == periodo_key:
            print(f"Sin cambios: período {periodo_key} ya procesado. Nada que hacer.")
            sys.exit(0)

    # 4. Consolidar instituciones
    # Unión de instituciones de débito y crédito (bancarias y SAEs)
    inst_map: dict[str, dict] = {}  # nombre_normalizado -> datos

    def get_or_create(name: str, is_prepago: bool = False) -> dict:
        key = clean_name(name)
        if key not in inst_map:
            inst_map[key] = {
                "nombre":  name,
                "tipo":    tipo_for(name, is_prepago),
                "debito":  0,
                "credito": 0,
                "prepago": 0,
            }
        return inst_map[key]

    # Débito (bancos + cooperativas con débito)
    for name, val in deb.items():
        if val > 0:
            get_or_create(name)["debito"] = val

    # Crédito = titulares + adicionales
    all_cred_names = set(ctit.keys()) | set(cadic.keys())
    for name in all_cred_names:
        t = ctit.get(name, 0)
        a = cadic.get(name, 0)
        total_cred = t + a
        if total_cred > 0:
            get_or_create(name)["credito"] = total_cred

    # Prepago no bancario (SAEs, emisores especializados)
    for name, val in prep.items():
        if val > 0:
            inst = get_or_create(name, is_prepago=True)
            inst["prepago"] = val
            # Si solo tiene prepago y no tiene débito/crédito, es no_bancario
            if inst["debito"] == 0 and inst["credito"] == 0:
                inst["tipo"] = "no_bancario"

    # 4b. Corrección: cooperativas reportan sus tarjetas de prepago (Dale, etc.)
    #     en la serie de débito de BEST — reclasificar debito→prepago para ellas.
    for inst in inst_map.values():
        if inst["tipo"] == "cooperativa" and inst["debito"] > 0:
            print(f"  [corrección] {inst['nombre']}: {inst['debito']:,} débito → prepago (producto de prepago reportado como débito en BEST)")
            inst["prepago"] += inst["debito"]
            inst["debito"] = 0

    # 5. Filtrar instituciones con datos reales y ordenar por total descendente
    instituciones = [
        v for v in inst_map.values()
        if v["debito"] + v["credito"] + v["prepago"] > 0
    ]
    instituciones.sort(key=lambda x: x["debito"] + x["credito"] + x["prepago"], reverse=True)

    # Generar códigos simples (basados en posición, sin datos de códigos reales disponibles)
    for i, inst in enumerate(instituciones):
        inst["codigo"] = str(i + 1).zfill(3)

    output = {
        "periodo":    periodo_str,
        "actualizado": datetime.utcnow().strftime("%Y-%m-%d"),
        "instituciones": instituciones,
    }

    print(f"\nInstituciones consolidadas: {len(instituciones)}")
    for inst in instituciones[:5]:
        print(f"  {inst['nombre']}: deb={inst['debito']:,} cred={inst['credito']:,} prep={inst['prepago']:,} [{inst['tipo']}]")

    # 6. Guardar JSON
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(output, ensure_ascii=False, indent=2))
    print(f"\nJSON guardado en {OUTPUT_JSON}")

    # 7. Guardar período procesado
    LAST_PERIOD.write_text(periodo_key)
    print(f"Período guardado en {LAST_PERIOD}")
    print("Listo.")


if __name__ == "__main__":
    main()
